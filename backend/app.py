from flask import Flask, request, jsonify, make_response
from flask_cors import CORS
import psycopg2
import psycopg2.extras
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
PHT = ZoneInfo("Asia/Manila")
import csv
import io
import hashlib
import os

app = Flask(__name__)
CORS(app, origins="*")

# ── Database ──────────────────────────────────────────────────────────────────
# Set DATABASE_URL in Vercel environment variables
# Format: postgresql://postgres:[PASSWORD]@db.[PROJECT-REF].supabase.co:5432/postgres

DATABASE_URL = os.environ.get("DATABASE_URL")

def get_db():
    conn = psycopg2.connect(DATABASE_URL)
    return conn

def hash_password(password: str) -> str:
    return hashlib.sha256(password.encode()).hexdigest()

# ── Helpers ────────────────────────────────────────────────────────────────────

def error(msg, code=400):
    return jsonify({"message": msg}), code

def require_fields(data, *fields):
    for f in fields:
        if not data.get(f) and data.get(f) != 0:
            return f"Missing required field: {f}"
    return None

def row_to_dict(row, cursor):
    """Convert psycopg2 row to dict using cursor description."""
    if row is None:
        return None
    columns = [desc[0] for desc in cursor.description]
    return dict(zip(columns, row))

def rows_to_list(rows, cursor):
    columns = [desc[0] for desc in cursor.description]
    return [dict(zip(columns, row)) for row in rows]

# ── Auth ──────────────────────────────────────────────────────────────────────

def log_activity(username, action, details=""):
    """Log admin activity to the database."""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO activity_logs (admin_username, action, details) VALUES (%s, %s, %s)",
            (username, action, details)
        )
        conn.commit()
        cur.close()
        conn.close()
    except Exception:
        pass  # Don't crash if logging fails


@app.route("/login", methods=["POST"])
def login():
    data = request.json or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    if not username or not password:
        return error("Username and password are required.")
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT username, role, status FROM admins WHERE username=%s AND password=%s",
        (username, hash_password(password))
    )
    row = cur.fetchone()
    cur.close()
    conn.close()
    if not row:
        return error("Invalid username or password.", 401)
    admin_username, role, status = row
    if status == "pending":
        return error("Your account is pending approval by the primary admin.", 403)
    if status == "disabled":
        return error("Your account has been disabled. Contact the primary admin.", 403)
    log_activity(admin_username, "LOGIN", f"Logged in")
    # Auto time-in: record shift start in admin_dtr
    try:
        today    = datetime.now(PHT).strftime("%Y-%m-%d")
        time_now = datetime.now(PHT).strftime("%Y-%m-%d %H:%M:%S")
        conn2 = get_db(); cur2 = conn2.cursor()
        # Close any unclosed shifts first (e.g. crash/browser close)
        cur2.execute("UPDATE admin_dtr SET time_out=%s WHERE admin_username=%s AND time_out IS NULL", (time_now, admin_username))
        # Open new shift
        cur2.execute("INSERT INTO admin_dtr (admin_username, date, time_in) VALUES (%s,%s,%s)",
            (admin_username, today, time_now))
        conn2.commit()
        shift_id = cur2.lastrowid if cur2.lastrowid else None
        # Get shift_id via SELECT since lastrowid may not work with psycopg2
        cur2.execute("SELECT id FROM admin_dtr WHERE admin_username=%s AND time_out IS NULL ORDER BY id DESC LIMIT 1", (admin_username,))
        row = cur2.fetchone()
        shift_id = row[0] if row else None
        cur2.close(); conn2.close()
    except Exception:
        shift_id = None
    return jsonify({"success": True, "username": admin_username, "role": role, "shift_id": shift_id})


@app.route("/register", methods=["POST"])
def register():
    data = request.json or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")
    if not username or not password:
        return error("Username and password are required.")
    if len(password) < 6:
        return error("Password must be at least 6 characters.")
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "INSERT INTO admins (username, password, role, status) VALUES (%s, %s, 'admin', 'pending')",
            (username, hash_password(password))
        )
        conn.commit()
        cur.close()
        conn.close()
        # Notify owner
        try:
            nc = get_db(); nr = nc.cursor()
            nr.execute(
                "INSERT INTO notifications (type, title, message) VALUES (%s, %s, %s)",
                ("NEW_ADMIN", "New Admin Request", f"{username} is requesting admin access.")
            )
            nc.commit(); nr.close(); nc.close()
        except Exception:
            pass
        return jsonify({"message": "Registration submitted! Please wait for the owner to approve your account."}), 201
    except psycopg2.errors.UniqueViolation:
        return error("Username already exists.", 409)


# ── Notifications ─────────────────────────────────────────────────────────────

@app.route("/notifications", methods=["GET"])
def get_notifications():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM notifications ORDER BY created_at DESC LIMIT 50")
    rows = rows_to_list(cur.fetchall(), cur)
    cur.close(); conn.close()
    return jsonify(rows)

@app.route("/notifications/unread-count", methods=["GET"])
def unread_count():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM notifications WHERE is_read=FALSE")
    count = cur.fetchone()[0]
    cur.close(); conn.close()
    return jsonify({"count": count})

@app.route("/notifications/mark-read", methods=["POST"])
def mark_notifications_read():
    conn = get_db(); cur = conn.cursor()
    cur.execute("UPDATE notifications SET is_read=TRUE WHERE is_read=FALSE")
    conn.commit(); cur.close(); conn.close()
    return jsonify({"message": "Marked all as read"})


# ── Superadmin: Manage Admins ─────────────────────────────────────────────────

@app.route("/admins", methods=["GET"])
def get_admins():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT username, role, status FROM admins ORDER BY role DESC, username ASC")
    rows = rows_to_list(cur.fetchall(), cur)
    cur.close()
    conn.close()
    return jsonify(rows)


@app.route("/admins/<username>/approve", methods=["POST"])
def approve_admin(username):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE admins SET status='active' WHERE username=%s AND role='admin'", (username,))
    conn.commit()
    affected = cur.rowcount
    cur.close()
    conn.close()
    if affected == 0:
        return error("Admin not found or cannot be modified.", 404)
    log_activity("superadmin", "APPROVE_ADMIN", f"Approved admin: {username}")
    return jsonify({"message": f"{username} approved successfully."})


@app.route("/admins/<username>/reject", methods=["POST"])
def reject_admin(username):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM admins WHERE username=%s AND role='admin' AND status='pending'", (username,))
    conn.commit()
    affected = cur.rowcount
    cur.close()
    conn.close()
    if affected == 0:
        return error("Admin not found.", 404)
    log_activity("superadmin", "REJECT_ADMIN", f"Rejected admin: {username}")
    return jsonify({"message": f"{username} rejected and removed."})


@app.route("/admins/<username>/disable", methods=["POST"])
def disable_admin(username):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE admins SET status='disabled' WHERE username=%s AND role='admin'", (username,))
    conn.commit()
    affected = cur.rowcount
    cur.close()
    conn.close()
    if affected == 0:
        return error("Admin not found or cannot be modified.", 404)
    log_activity("superadmin", "DISABLE_ADMIN", f"Disabled admin: {username}")
    return jsonify({"message": f"{username} has been disabled."})


@app.route("/admins/<username>/enable", methods=["POST"])
def enable_admin(username):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("UPDATE admins SET status='active' WHERE username=%s AND role='admin'", (username,))
    conn.commit()
    cur.close()
    conn.close()
    log_activity("superadmin", "ENABLE_ADMIN", f"Enabled admin: {username}")
    return jsonify({"message": f"{username} has been enabled."})


@app.route("/admins/<username>", methods=["DELETE"])
def delete_admin(username):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM admins WHERE username=%s AND role='admin'", (username,))
    conn.commit()
    affected = cur.rowcount
    cur.close()
    conn.close()
    if affected == 0:
        return error("Admin not found or cannot be deleted.", 404)
    log_activity("superadmin", "DELETE_ADMIN", f"Deleted admin: {username}")
    return jsonify({"message": f"{username} deleted."})


@app.route("/admins/pending/count", methods=["GET"])
def pending_count():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM admins WHERE status='pending'")
    count = cur.fetchone()[0]
    cur.close()
    conn.close()
    return jsonify({"count": count})


# ── Activity Logs ─────────────────────────────────────────────────────────────

@app.route("/logs", methods=["GET"])
def get_logs():
    limit    = int(request.args.get("limit", 100))
    username = request.args.get("username", "")
    conn = get_db()
    cur = conn.cursor()
    if username:
        cur.execute(
            "SELECT * FROM activity_logs WHERE admin_username=%s ORDER BY created_at DESC LIMIT %s",
            (username, limit)
        )
    else:
        cur.execute(
            "SELECT * FROM activity_logs ORDER BY created_at DESC LIMIT %s",
            (limit,)
        )
    rows = rows_to_list(cur.fetchall(), cur)
    cur.close()
    conn.close()
    return jsonify(rows)

# ── Members ───────────────────────────────────────────────────────────────────

@app.route("/members", methods=["GET"])
def get_members():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM members ORDER BY id DESC")
    rows = rows_to_list(cur.fetchall(), cur)
    cur.close()
    conn.close()
    return jsonify(rows)

@app.route("/members/<int:member_id>", methods=["GET"])
def get_member(member_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM members WHERE id=%s", (member_id,))
    row = row_to_dict(cur.fetchone(), cur)
    cur.close()
    conn.close()
    if row:
        return jsonify(row)
    return error("Member not found.", 404)

@app.route("/members", methods=["POST"])
def add_member():
    data = request.json or {}
    err = require_fields(data, "name", "months", "price")
    if err:
        return error(err)
    try:
        months = int(data["months"])
        price = float(data["price"])
        discount = float(data.get("discount", 0))
        if months < 1 or months > 60:
            return error("Months must be between 1 and 60.")
        if price < 0:
            return error("Price cannot be negative.")
        if not (0 <= discount <= 100):
            return error("Discount must be between 0 and 100.")
    except (ValueError, TypeError):
        return error("Invalid numeric values.")

    start = datetime.now(PHT)
    expiration = start + timedelta(days=30 * months)
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO members (name, email, phone, plan, months, price, discount, start_date, expiration_date, status)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, 'active')
    """, (
        data["name"].strip(),
        data.get("email", "").strip(),
        data.get("phone", "").strip(),
        data.get("plan", "Basic"),
        months, price, discount,
        start.strftime("%Y-%m-%d"),
        expiration.strftime("%Y-%m-%d")
    ))
    conn.commit()
    cur.close()
    conn.close()
    admin_user = request.headers.get("X-Admin-User", "unknown")
    log_activity(admin_user, "ADD_MEMBER", f"Added member: {data.get('name','')}")
    return jsonify({"message": "Member added"}), 201

@app.route("/members/<int:member_id>", methods=["PUT"])
def update_member(member_id):
    data = request.json or {}
    err = require_fields(data, "name", "months", "price")
    if err:
        return error(err)
    try:
        months = int(data["months"])
        price = float(data["price"])
        discount = float(data.get("discount", 0))
    except (ValueError, TypeError):
        return error("Invalid numeric values.")

    start_str = data.get("start_date") or datetime.now(PHT).strftime("%Y-%m-%d")
    try:
        expiration = (datetime.strptime(start_str, "%Y-%m-%d") + timedelta(days=30 * months)).strftime("%Y-%m-%d")
    except ValueError:
        return error("Invalid start_date format. Use YYYY-MM-DD.")

    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE members SET name=%s, email=%s, phone=%s, plan=%s, months=%s, price=%s,
        discount=%s, start_date=%s, expiration_date=%s WHERE id=%s
    """, (
        data["name"].strip(),
        data.get("email", "").strip(),
        data.get("phone", "").strip(),
        data.get("plan", "Basic"),
        months, price, discount,
        start_str, expiration, member_id
    ))
    conn.commit()
    rows_affected = cur.rowcount
    cur.close()
    conn.close()
    if rows_affected == 0:
        return error("Member not found.", 404)
    admin_user = request.headers.get("X-Admin-User", "unknown")
    log_activity(admin_user, "EDIT_MEMBER", f"Updated member ID: {member_id}")
    return jsonify({"message": "Member updated"})

@app.route("/members/<int:member_id>", methods=["DELETE"])
def delete_member(member_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM members WHERE id=%s", (member_id,))
    conn.commit()
    rows_affected = cur.rowcount
    cur.close()
    conn.close()
    if rows_affected == 0:
        return error("Member not found.", 404)
    admin_user = request.headers.get("X-Admin-User", "unknown")
    log_activity(admin_user, "DELETE_MEMBER", f"Deleted member ID: {member_id}")
    return jsonify({"message": "Deleted"})

# ── Attendance ────────────────────────────────────────────────────────────────

@app.route("/attendance/timein", methods=["POST"])
def time_in():
    data = request.json or {}
    member_id = data.get("member_id")
    if not member_id:
        return error("member_id is required.")
    now = datetime.now(PHT)
    today = now.strftime("%Y-%m-%d")
    time_now = now.strftime("%I:%M %p")
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT id FROM attendance WHERE member_id=%s AND date=%s AND time_out IS NULL",
        (member_id, today)
    )
    if cur.fetchone():
        cur.close()
        conn.close()
        return error("Member is already timed in today.", 409)
    cur.execute("SELECT name FROM members WHERE id=%s", (member_id,))
    member = cur.fetchone()
    if not member:
        cur.close()
        conn.close()
        return error("Member not found.", 404)
    cur.execute(
        "INSERT INTO attendance (member_id, member_name, time_in, date) VALUES (%s, %s, %s, %s)",
        (member_id, member[0], time_now, today)
    )
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"message": f"{member[0]} timed in at {time_now}"}), 201

@app.route("/attendance/timeout", methods=["POST"])
def time_out():
    data = request.json or {}
    member_id = data.get("member_id")
    if not member_id:
        return error("member_id is required.")
    now = datetime.now(PHT)
    today = now.strftime("%Y-%m-%d")
    time_now = now.strftime("%I:%M %p")
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT id, member_name FROM attendance WHERE member_id=%s AND date=%s AND time_out IS NULL",
        (member_id, today)
    )
    record = cur.fetchone()
    if not record:
        cur.close()
        conn.close()
        return error("No active time-in found for this member today.", 404)
    cur.execute("UPDATE attendance SET time_out=%s WHERE id=%s", (time_now, record[0]))
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"message": f"{record[1]} timed out at {time_now}"})

@app.route("/attendance", methods=["GET"])
def get_attendance():
    date = request.args.get("date", datetime.now(PHT).strftime("%Y-%m-%d"))
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM attendance WHERE date=%s ORDER BY id DESC", (date,))
    rows = rows_to_list(cur.fetchall(), cur)
    cur.execute(
        "SELECT COUNT(*) FROM attendance WHERE date=%s AND time_out IS NULL", (date,)
    )
    still_in = cur.fetchone()[0]
    cur.close()
    conn.close()
    return jsonify({
        "records": rows,
        "total": len(rows),
        "still_in": still_in,
        "date": date
    })

# ── Walk-ins ──────────────────────────────────────────────────────────────────

@app.route("/walkins", methods=["GET"])
def get_walkins():
    date = request.args.get("date", datetime.now(PHT).strftime("%Y-%m-%d"))
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM walkins WHERE date=%s ORDER BY id DESC", (date,))
    rows = rows_to_list(cur.fetchall(), cur)
    cur.execute("SELECT COALESCE(SUM(amount), 0) FROM walkins WHERE date=%s", (date,))
    total = cur.fetchone()[0]
    cur.close()
    conn.close()
    return jsonify({"walkins": rows, "total": round(float(total), 2), "date": date})

@app.route("/walkins", methods=["POST"])
def add_walkin():
    data = request.json or {}
    err = require_fields(data, "name", "amount")
    if err:
        return error(err)
    try:
        amount = float(data["amount"])
        if amount <= 0:
            return error("Amount must be greater than 0.")
    except (ValueError, TypeError):
        return error("Invalid amount.")
    date = data.get("date") or datetime.now(PHT).strftime("%Y-%m-%d")
    conn = get_db()
    cur = conn.cursor()
    now_ts = datetime.now(PHT).strftime("%Y-%m-%d %H:%M:%S")
    cur.execute(
        "INSERT INTO walkins (name, amount, note, date, created_at) VALUES (%s, %s, %s, %s, %s)",
        (data["name"].strip(), amount, data.get("note", "").strip(), date, now_ts)
    )
    conn.commit()
    cur.close()
    conn.close()
    return jsonify({"message": "Walk-in recorded"}), 201

@app.route("/walkins/<int:walkin_id>", methods=["DELETE"])
def delete_walkin(walkin_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM walkins WHERE id=%s", (walkin_id,))
    conn.commit()
    rows_affected = cur.rowcount
    cur.close()
    conn.close()
    if rows_affected == 0:
        return error("Walk-in not found.", 404)
    return jsonify({"message": "Deleted"})

# ── Stats ─────────────────────────────────────────────────────────────────────

@app.route("/stats", methods=["GET"])
def stats():
    today = datetime.now(PHT).strftime("%Y-%m-%d")
    conn = get_db()
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*) FROM members")
    total = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM members WHERE expiration_date >= TO_CHAR(CURRENT_DATE, 'YYYY-MM-DD')")
    active = cur.fetchone()[0]

    cur.execute("SELECT COALESCE(SUM(price - (price * discount / 100)), 0) FROM members")
    revenue = cur.fetchone()[0]

    cur.execute(
        "SELECT COUNT(*) FROM members WHERE TO_CHAR(start_date::date, 'YYYY-MM') = TO_CHAR(CURRENT_DATE, 'YYYY-MM')"
    )
    new_this_month = cur.fetchone()[0]

    # Shift-scoped walk-in revenue: only count walkins from this admin's current shift login time
    shift_id = request.args.get("shift_id")
    shift_start = None
    if shift_id:
        cur.execute("SELECT time_in FROM admin_dtr WHERE id=%s AND time_out IS NULL", (shift_id,))
        row = cur.fetchone()
        if row:
            shift_start = row[0]
    if shift_start:
        cur.execute("SELECT COALESCE(SUM(amount),0) FROM walkins WHERE created_at >= %s", (shift_start,))
        walkin_today = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM walkins WHERE created_at >= %s", (shift_start,))
        walkin_count = cur.fetchone()[0]
    else:
        cur.execute("SELECT COALESCE(SUM(amount), 0) FROM walkins WHERE date=%s", (today,))
        walkin_today = cur.fetchone()[0]
        cur.execute("SELECT COUNT(*) FROM walkins WHERE date=%s", (today,))
        walkin_count = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM attendance WHERE date=%s AND time_out IS NULL", (today,))
    members_in_gym = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM attendance WHERE date=%s", (today,))
    visits_today = cur.fetchone()[0]

    cur.close()
    conn.close()

    return jsonify({
        "total_members": total,
        "active_members": active,
        "revenue": round(float(revenue), 2),
        "new_this_month": new_this_month,
        "walkin_revenue_today": round(float(walkin_today), 2),
        "walkin_count_today": walkin_count,
        "members_in_gym": members_in_gym,
        "visits_today": visits_today
    })

# ── Expiring ──────────────────────────────────────────────────────────────────

@app.route("/expiring", methods=["GET"])
def expiring():
    try:
        days = max(1, min(int(request.args.get("days", 7)), 365))
    except ValueError:
        days = 7
    today = datetime.now(PHT).strftime("%Y-%m-%d")
    future = (datetime.now(PHT) + timedelta(days=days)).strftime("%Y-%m-%d")
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM members WHERE expiration_date <= %s AND expiration_date >= %s ORDER BY expiration_date ASC",
        (future, today)
    )
    expiring_soon = rows_to_list(cur.fetchall(), cur)
    cur.execute(
        "SELECT * FROM members WHERE expiration_date < %s ORDER BY expiration_date DESC",
        (today,)
    )
    expired = rows_to_list(cur.fetchall(), cur)
    cur.close()
    conn.close()
    return jsonify({"expiring_soon": expiring_soon, "expired": expired})

# ── Export CSV ────────────────────────────────────────────────────────────────

@app.route("/export/csv", methods=["GET"])
def export_csv():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT * FROM members ORDER BY id DESC")
    rows = rows_to_list(cur.fetchall(), cur)
    cur.close()
    conn.close()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["ID", "Name", "Email", "Phone", "Plan", "Months",
                     "Price", "Discount%", "Net Price", "Start Date", "Expiration Date", "Status"])
    today_str = datetime.now(PHT).strftime("%Y-%m-%d")
    for r in rows:
        try:
            net = float(r["price"] or 0) - (float(r["price"] or 0) * float(r["discount"] or 0) / 100)
            exp = r["expiration_date"] or ""
            if not exp or exp < today_str:
                status = "Expired"
            else:
                days_left = (datetime.strptime(exp, "%Y-%m-%d") - datetime.now(PHT).replace(tzinfo=None)).days
                status = "Expiring Soon" if days_left <= 7 else "Active"
        except Exception:
            net = 0; status = "Unknown"
        writer.writerow([r["id"], r["name"], r["email"], r["phone"], r["plan"],
                         r["months"], r["price"], r["discount"], round(net, 2),
                         r["start_date"], r["expiration_date"], status])
    response = make_response(output.getvalue())
    response.headers["Content-Disposition"] = "attachment; filename=members.csv"
    response.headers["Content-Type"] = "text/csv; charset=utf-8"
    return response

# ── Monthly Excel Report ──────────────────────────────────────────────────────

@app.route("/report/excel", methods=["GET"])
def report_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return error("openpyxl not installed.", 500)
    try:

        year = request.args.get("year", datetime.now(PHT).strftime("%Y"))
        month = request.args.get("month", datetime.now(PHT).strftime("%m")).zfill(2)
        month_str = f"{year}-{month}"
        try:
            month_name = datetime.strptime(month_str, "%Y-%m").strftime("%B %Y")
        except ValueError:
            return error("Invalid year/month.")

        conn = get_db()
        cur = conn.cursor()

        cur.execute("SELECT * FROM members WHERE TO_CHAR(start_date::date,'YYYY-MM')=%s", (month_str,))
        new_members = rows_to_list(cur.fetchall(), cur)

        cur.execute("SELECT * FROM members ORDER BY name ASC")
        all_members = rows_to_list(cur.fetchall(), cur)

        cur.execute("SELECT * FROM members WHERE expiration_date >= TO_CHAR(CURRENT_DATE, 'YYYY-MM-DD')")
        active_members = rows_to_list(cur.fetchall(), cur)

        cur.execute("SELECT * FROM walkins WHERE TO_CHAR(date::date,'YYYY-MM')=%s ORDER BY date ASC", (month_str,))
        walkins = rows_to_list(cur.fetchall(), cur)

        cur.execute("""
            SELECT a.*, m.plan FROM attendance a
            LEFT JOIN members m ON a.member_id = m.id
            WHERE TO_CHAR(a.date::date,'YYYY-MM') = %s
            ORDER BY a.date ASC, a.time_in ASC
        """, (month_str,))
        attendance = rows_to_list(cur.fetchall(), cur)
        cur.close()
        conn.close()

        member_revenue = sum((r["price"] - r["price"] * r["discount"] / 100) for r in new_members)
        walkin_revenue = sum(r["amount"] for r in walkins)
        total_revenue = member_revenue + walkin_revenue

        DARK = "1a1a1a"; YELLOW = "E8FF00"; HEADER_BG = "2a2a2a"
        WHITE = "F0F0F0"; GREEN = "00E676"; MUTED = "888888"; ROW_ALT = "1f1f1f"
        thin = Side(style="thin", color="2a2a2a")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        wb = openpyxl.Workbook()

        def style_header_row(ws, row, cols, bg=HEADER_BG, fg=YELLOW):
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(bold=True, color=fg, name="Arial", size=10)
                cell.fill = PatternFill("solid", start_color=bg)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

        def style_data_row(ws, row, cols, alt=False):
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = Font(color=WHITE, name="Arial", size=9)
                cell.fill = PatternFill("solid", start_color=ROW_ALT if alt else DARK)
                cell.alignment = Alignment(vertical="center")
                cell.border = border

        def fill_bg(ws, max_row, max_col):
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.fill = PatternFill("solid", start_color=DARK)

        def set_col_widths(ws, widths):
            for i, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        def sheet_title(ws, row, cols, text):
            ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
            c = ws[f"A{row}"]
            c.value = text
            c.font = Font(bold=True, color=YELLOW, name="Arial", size=14)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", start_color=DARK)
            ws.row_dimensions[row].height = 28

        ws1 = wb.active
        ws1.title = "Summary"
        ws1.sheet_view.showGridLines = False
        fill_bg(ws1, 60, 10)
        ws1.merge_cells("A1:H1")
        c = ws1["A1"]
        c.value = "LOYD'S FITNESS GYM"
        c.font = Font(bold=True, color=YELLOW, name="Arial", size=22)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", start_color=DARK)
        ws1.row_dimensions[1].height = 40
        ws1.merge_cells("A2:H2")
        c = ws1["A2"]
        c.value = f"Monthly Report — {month_name}"
        c.font = Font(color=WHITE, name="Arial", size=13)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = PatternFill("solid", start_color=DARK)
        ws1.row_dimensions[2].height = 24

        kpis = [
            ("TOTAL MEMBERS", len(all_members), WHITE),
            ("ACTIVE MEMBERS", len(active_members), GREEN),
            ("NEW THIS MONTH", len(new_members), YELLOW),
            ("WALK-INS", len(walkins), "00B0FF"),
            ("ATTENDANCE LOGS", len(attendance), "FF9100"),
            ("MEMBER REVENUE", f"₱{member_revenue:,.2f}", GREEN),
            ("WALKIN REVENUE", f"₱{walkin_revenue:,.2f}", YELLOW),
            ("TOTAL REVENUE", f"₱{total_revenue:,.2f}", "FF4D00"),
        ]
        for col, (label, value, color) in enumerate(kpis, 1):
            lc = ws1.cell(row=5, column=col, value=label)
            lc.font = Font(color=MUTED, name="Arial", size=8, bold=True)
            lc.fill = PatternFill("solid", start_color=HEADER_BG)
            lc.alignment = Alignment(horizontal="center", vertical="center")
            lc.border = border
            vc = ws1.cell(row=6, column=col, value=value)
            vc.font = Font(bold=True, color=color, name="Arial", size=14)
            vc.fill = PatternFill("solid", start_color=HEADER_BG)
            vc.alignment = Alignment(horizontal="center", vertical="center")
            vc.border = border
        set_col_widths(ws1, [4, 20, 22, 14, 12, 8, 12, 10, 12, 12])

        ws2 = wb.create_sheet("All Members")
        ws2.sheet_view.showGridLines = False
        fill_bg(ws2, len(all_members) + 10, 11)
        sheet_title(ws2, 1, 11, f"ALL MEMBERS — {month_name}")
        h2 = ["#", "Name", "Email", "Phone", "Plan", "Months", "Net Price (₱)", "Discount %", "Start Date", "Expiration", "Status"]
        for i, h in enumerate(h2, 1):
            ws2.cell(row=2, column=i, value=h)
        style_header_row(ws2, 2, len(h2))
        today_str = datetime.now(PHT).strftime("%Y-%m-%d")
        for idx, m in enumerate(all_members):
            r = 3 + idx
            net = m["price"] - (m["price"] * m["discount"] / 100)
            exp = m["expiration_date"]
            days_left = (datetime.strptime(exp, "%Y-%m-%d") - datetime.now(PHT)).days
            if exp < today_str: status, color = "Expired", "FF1744"
            elif days_left <= 7: status, color = "Expiring Soon", "FFB300"
            else: status, color = "Active", GREEN
            for c, val in enumerate([idx+1, m["name"], m["email"], m["phone"], m["plan"],
                                      m["months"], round(net, 2), f'{m["discount"]}%',
                                      m["start_date"], exp, status], 1):
                ws2.cell(row=r, column=c, value=val)
            style_data_row(ws2, r, len(h2), alt=idx % 2 == 1)
            ws2.cell(row=r, column=11).font = Font(color=color, name="Arial", size=9, bold=True)
        set_col_widths(ws2, [4, 20, 22, 14, 12, 8, 14, 10, 12, 12, 14])

        ws3 = wb.create_sheet("Walk-ins")
        ws3.sheet_view.showGridLines = False
        fill_bg(ws3, len(walkins) + 10, 5)
        sheet_title(ws3, 1, 5, f"WALK-IN REVENUE — {month_name}")
        h3 = ["#", "Name", "Amount (₱)", "Note", "Date"]
        for i, h in enumerate(h3, 1):
            ws3.cell(row=2, column=i, value=h)
        style_header_row(ws3, 2, len(h3))
        for idx, w in enumerate(walkins):
            r = 3 + idx
            for c, val in enumerate([idx+1, w["name"], w["amount"], w["note"] or "—", w["date"]], 1):
                ws3.cell(row=r, column=c, value=val)
            style_data_row(ws3, r, len(h3), alt=idx % 2 == 1)
        set_col_widths(ws3, [4, 22, 14, 24, 14])

        ws4 = wb.create_sheet("Attendance")
        ws4.sheet_view.showGridLines = False
        fill_bg(ws4, len(attendance) + 10, 7)
        sheet_title(ws4, 1, 7, f"ATTENDANCE LOG — {month_name}")
        h4 = ["#", "Member", "Plan", "Date", "Time In", "Time Out", "Status"]
        for i, h in enumerate(h4, 1):
            ws4.cell(row=2, column=i, value=h)
        style_header_row(ws4, 2, len(h4))
        for idx, a in enumerate(attendance):
            r = 3 + idx
            status = "Inside" if not a["time_out"] else "Left"
            for c, val in enumerate([idx+1, a["member_name"], a.get("plan") or "—",
                                      a["date"], a["time_in"], a["time_out"] or "—", status], 1):
                ws4.cell(row=r, column=c, value=val)
            style_data_row(ws4, r, len(h4), alt=idx % 2 == 1)
            ws4.cell(row=r, column=7).font = Font(color=GREEN if status == "Inside" else "FF1744", name="Arial", size=9, bold=True)
        set_col_widths(ws4, [4, 22, 14, 12, 12, 12, 10])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        filename = f"LoydsGym_Report_{month_name.replace(' ', '_')}.xlsx"
        response = make_response(output.getvalue())
        response.headers["Content-Disposition"] = f"attachment; filename={filename}"
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return response
    except Exception as ex:
        return error(f"Error generating report: {str(ex)}", 500)


    except Exception as ex:
        return error(f"Error generating report: {str(ex)}", 500)



@app.route("/shifts/daily", methods=["GET"])
def shifts_daily():
    date = request.args.get("date", datetime.now(PHT).strftime("%Y-%m-%d"))
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT id, admin_username, time_in, time_out, shift_revenue,
                   time_in AS shift_ts_in, time_out AS shift_ts_out,
                   date
            FROM admin_dtr
            WHERE date = %s
            ORDER BY time_in DESC NULLS LAST
        """, (date,))
        shifts = rows_to_list(cur.fetchall(), cur)
        return jsonify({"shifts": shifts})
    except Exception as ex:
        return error(f"Error fetching daily shifts: {str(ex)}", 500)
    finally:
        cur.close()
        conn.close()

@app.route("/shifts", methods=["GET"])
def get_shifts():
    month = request.args.get("month", datetime.now(PHT).strftime("%m"))
    year  = request.args.get("year",  datetime.now(PHT).strftime("%Y"))
    admin = request.args.get("admin", "")
    prefix = f"{year}-{month.zfill(2)}"
    conn = get_db(); cur = conn.cursor()
    if admin:
        cur.execute("""
            SELECT id, admin_username, date, time_in, time_out,
                   COALESCE(shift_revenue,0) as shift_revenue
            FROM admin_dtr
            WHERE date LIKE %s AND admin_username=%s
            ORDER BY time_in DESC
        """, (f"{prefix}%", admin))
    else:
        cur.execute("""
            SELECT id, admin_username, date, time_in, time_out,
                   COALESCE(shift_revenue,0) as shift_revenue
            FROM admin_dtr
            WHERE date LIKE %s
            ORDER BY time_in DESC
        """, (f"{prefix}%",))
    rows = rows_to_list(cur.fetchall(), cur)
    # Summary per admin
    summary = {}
    for r in rows:
        u = r["admin_username"]
        if u not in summary:
            summary[u] = {"admin": u, "total_shifts": 0, "total_revenue": 0.0, "total_hours": 0.0}
        summary[u]["total_shifts"] += 1
        summary[u]["total_revenue"] += float(r["shift_revenue"] or 0)
        if r["time_in"] and r["time_out"]:
            try:
                ti = datetime.strptime(str(r["time_in"])[:19], "%Y-%m-%d %H:%M:%S")
                to = datetime.strptime(str(r["time_out"])[:19], "%Y-%m-%d %H:%M:%S")
                summary[u]["total_hours"] += round((to - ti).seconds / 3600, 2)
            except Exception:
                pass
    cur.close(); conn.close()
    return jsonify({"shifts": rows, "summary": list(summary.values())})


# ── Admin DTR ─────────────────────────────────────────────────────────────────

@app.route("/admin/dtr/timein", methods=["POST"])
def admin_time_in():
    data = request.json or {}
    username = data.get("username", "").strip()
    if not username:
        return error("Username is required.")
    today = datetime.now(PHT).strftime("%Y-%m-%d")
    time_now = datetime.now(PHT).strftime("%I:%M %p")
    conn = get_db()
    cur = conn.cursor()
    # Check if already timed in today without timing out
    cur.execute(
        "SELECT id FROM admin_dtr WHERE admin_username=%s AND date=%s AND time_out IS NULL",
        (username, today)
    )
    existing = cur.fetchone()
    if existing:
        cur.close(); conn.close()
        return error(f"{username} is already timed in today.")
    cur.execute(
        "INSERT INTO admin_dtr (admin_username, date, time_in) VALUES (%s, %s, %s)",
        (username, today, time_now)
    )
    conn.commit()
    cur.close(); conn.close()
    return jsonify({"message": f"Time in recorded for {username} at {time_now}"})


@app.route("/admin/dtr/timeout", methods=["POST"])
def admin_time_out():
    data = request.json or {}
    username = data.get("username", "").strip()
    if not username:
        return error("Username is required.")
    today = datetime.now(PHT).strftime("%Y-%m-%d")
    time_now = datetime.now(PHT).strftime("%I:%M %p")
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT id FROM admin_dtr WHERE admin_username=%s AND date=%s AND time_out IS NULL",
        (username, today)
    )
    record = cur.fetchone()
    if not record:
        cur.close(); conn.close()
        return error(f"No active time-in found for {username} today.")
    cur.execute(
        "UPDATE admin_dtr SET time_out=%s WHERE id=%s",
        (time_now, record[0])
    )
    conn.commit()
    cur.close(); conn.close()
    return jsonify({"message": f"Time out recorded for {username} at {time_now}"})


@app.route("/admin/dtr", methods=["GET"])
def get_admin_dtr():
    date = request.args.get("date", datetime.now(PHT).strftime("%Y-%m-%d"))
    username = request.args.get("username", "")
    conn = get_db()
    cur = conn.cursor()
    if username:
        cur.execute(
            "SELECT * FROM admin_dtr WHERE admin_username=%s ORDER BY date DESC, id DESC",
            (username,)
        )
    else:
        cur.execute(
            "SELECT * FROM admin_dtr WHERE date=%s ORDER BY id DESC",
            (date,)
        )
    rows = rows_to_list(cur.fetchall(), cur)
    cur.close(); conn.close()
    return jsonify(rows)


@app.route("/admin/dtr/all", methods=["GET"])
def get_all_admin_dtr():
    month = request.args.get("month", datetime.now(PHT).strftime("%m"))
    year  = request.args.get("year",  datetime.now(PHT).strftime("%Y"))
    prefix = f"{year}-{month.zfill(2)}"
    conn = get_db()
    cur = conn.cursor()
    cur.execute(
        "SELECT * FROM admin_dtr WHERE date LIKE %s ORDER BY date DESC, admin_username ASC",
        (f"{prefix}%",)
    )
    rows = rows_to_list(cur.fetchall(), cur)
    cur.close(); conn.close()
    return jsonify(rows)


@app.route("/admin/dtr/<int:dtr_id>", methods=["DELETE"])
def delete_admin_dtr(dtr_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM admin_dtr WHERE id=%s", (dtr_id,))
    conn.commit()
    affected = cur.rowcount
    cur.close(); conn.close()
    if affected == 0:
        return error("Record not found.", 404)
    return jsonify({"message": "Record deleted"})


@app.route("/logout", methods=["POST"])
def logout():
    data     = request.json or {}
    username = data.get("username", "").strip()
    shift_id = data.get("shift_id")
    if not username:
        return jsonify({"message": "Logged out"})
    now_str = datetime.now(PHT).strftime("%Y-%m-%d %H:%M:%S")
    conn = get_db(); cur = conn.cursor()
    # Find the open shift
    if shift_id:
        cur.execute("SELECT id, time_in FROM admin_dtr WHERE id=%s AND time_out IS NULL", (shift_id,))
    else:
        cur.execute("SELECT id, time_in FROM admin_dtr WHERE admin_username=%s AND time_out IS NULL ORDER BY id DESC LIMIT 1", (username,))
    record = cur.fetchone()
    shift_revenue = 0
    if record:
        open_id   = record[0]
        time_in   = record[1]
        # Calculate walk-in revenue collected during this shift
        try:
            cur.execute("""
                SELECT COALESCE(SUM(amount),0) FROM walkins
                WHERE created_at >= %s AND created_at <= %s
            """, (time_in, now_str))
            shift_revenue = float(cur.fetchone()[0] or 0)
        except Exception:
            shift_revenue = 0
        cur.execute("UPDATE admin_dtr SET time_out=%s, shift_revenue=%s WHERE id=%s",
            (now_str, shift_revenue, open_id))
        conn.commit()
    cur.close(); conn.close()
    log_activity(username, "LOGOUT", f"Admin logged out. Shift revenue: P{shift_revenue:,.2f}")
    return jsonify({"message": "Logged out successfully", "shift_revenue": shift_revenue})


# ── Employee DTR ──────────────────────────────────────────────────────────────

@app.route("/employee/dtr/timein", methods=["POST"])
def employee_time_in():
    data        = request.json or {}
    name        = data.get("name", "").strip()
    note        = data.get("note", "").strip()
    recorded_by = request.headers.get("X-Admin-User", "unknown")
    if not name:
        return error("Employee name is required.")
    today    = datetime.now(PHT).strftime("%Y-%m-%d")
    time_now = datetime.now(PHT).strftime("%I:%M %p")
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT id FROM employee_dtr WHERE employee_name=%s AND date=%s AND time_out IS NULL", (name, today))
    if cur.fetchone():
        cur.close(); conn.close()
        return error(f"{name} is already timed in today.")
    cur.execute("INSERT INTO employee_dtr (employee_name, date, time_in, note, recorded_by) VALUES (%s,%s,%s,%s,%s)", (name, today, time_now, note, recorded_by))
    conn.commit(); cur.close(); conn.close()
    log_activity(recorded_by, "EMPLOYEE_TIMEIN", f"Time in for: {name}")
    return jsonify({"message": f"Time in recorded for {name} at {time_now}"})


@app.route("/employee/dtr/timeout", methods=["POST"])
def employee_time_out():
    data        = request.json or {}
    name        = data.get("name", "").strip()
    dtr_id      = data.get("id")
    recorded_by = request.headers.get("X-Admin-User", "unknown")
    today    = datetime.now(PHT).strftime("%Y-%m-%d")
    time_now = datetime.now(PHT).strftime("%I:%M %p")
    conn = get_db(); cur = conn.cursor()
    if dtr_id:
        cur.execute("SELECT id FROM employee_dtr WHERE id=%s AND time_out IS NULL", (dtr_id,))
    else:
        cur.execute("SELECT id FROM employee_dtr WHERE employee_name=%s AND date=%s AND time_out IS NULL", (name, today))
    record = cur.fetchone()
    if not record:
        cur.close(); conn.close()
        return error(f"No active time-in found for {name}.")
    cur.execute("UPDATE employee_dtr SET time_out=%s WHERE id=%s", (time_now, record[0]))
    conn.commit(); cur.close(); conn.close()
    log_activity(recorded_by, "EMPLOYEE_TIMEOUT", f"Time out for: {name}")
    return jsonify({"message": f"Time out recorded for {name} at {time_now}"})


@app.route("/employee/dtr", methods=["GET"])
def get_employee_dtr():
    date = request.args.get("date", datetime.now(PHT).strftime("%Y-%m-%d"))
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT * FROM employee_dtr WHERE date=%s ORDER BY id DESC", (date,))
    rows = rows_to_list(cur.fetchall(), cur)
    cur.close(); conn.close()
    return jsonify(rows)


@app.route("/employee/dtr/all", methods=["GET"])
def get_all_employee_dtr():
    month  = request.args.get("month", datetime.now(PHT).strftime("%m"))
    year   = request.args.get("year",  datetime.now(PHT).strftime("%Y"))
    name   = request.args.get("name", "")
    prefix = f"{year}-{month.zfill(2)}"
    conn = get_db(); cur = conn.cursor()
    if name:
        cur.execute("SELECT * FROM employee_dtr WHERE date LIKE %s AND employee_name=%s ORDER BY date DESC, id DESC", (f"{prefix}%", name))
    else:
        cur.execute("SELECT * FROM employee_dtr WHERE date LIKE %s ORDER BY date DESC, employee_name ASC", (f"{prefix}%",))
    rows = rows_to_list(cur.fetchall(), cur)
    cur.close(); conn.close()
    return jsonify(rows)


@app.route("/employee/dtr/employees", methods=["GET"])
def get_employee_names():
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT DISTINCT employee_name FROM employee_dtr ORDER BY employee_name ASC")
    names = [row[0] for row in cur.fetchall()]
    cur.close(); conn.close()
    return jsonify(names)


@app.route("/employee/dtr/<int:dtr_id>", methods=["DELETE"])
def delete_employee_dtr(dtr_id):
    conn = get_db(); cur = conn.cursor()
    cur.execute("DELETE FROM employee_dtr WHERE id=%s", (dtr_id,))
    conn.commit()
    affected = cur.rowcount
    cur.close(); conn.close()
    if affected == 0:
        return error("Record not found.", 404)
    return jsonify({"message": "Record deleted"})



@app.route("/owner/change-password", methods=["POST"])
def owner_change_password():
    data = request.json or {}
    current_pw  = data.get("current_password", "")
    new_pw      = data.get("new_password", "")
    if not current_pw or not new_pw:
        return error("All fields are required.")
    if len(new_pw) < 6:
        return error("New password must be at least 6 characters.")
    conn = get_db(); cur = conn.cursor()
    hashed_current = hashlib.sha256(current_pw.encode()).hexdigest()
    cur.execute("SELECT username FROM admins WHERE username='owner' AND password=%s", (hashed_current,))
    if not cur.fetchone():
        cur.close(); conn.close()
        return error("Current password is incorrect.")
    hashed_new = hashlib.sha256(new_pw.encode()).hexdigest()
    cur.execute("UPDATE admins SET password=%s WHERE username='owner'", (hashed_new,))
    conn.commit(); cur.close(); conn.close()
    log_activity("owner", "CHANGE_PASSWORD", "Owner changed their password")
    return jsonify({"message": "Password changed successfully!"})


@app.route("/admins/<username>/change-password", methods=["POST"])
def owner_change_admin_password(username):
    data   = request.json or {}
    new_pw = data.get("new_password", "").strip()
    if not new_pw:
        return error("New password is required.")
    if len(new_pw) < 6:
        return error("Password must be at least 6 characters.")
    conn = get_db(); cur = conn.cursor()
    cur.execute("SELECT username, role FROM admins WHERE username=%s", (username,))
    admin = cur.fetchone()
    if not admin:
        cur.close(); conn.close()
        return error("Admin not found.", 404)
    if admin[1] == "owner":
        cur.close(); conn.close()
        return error("Cannot change owner password from here.")
    hashed = hashlib.sha256(new_pw.encode()).hexdigest()
    cur.execute("UPDATE admins SET password=%s WHERE username=%s", (hashed, username))
    conn.commit(); cur.close(); conn.close()
    recorded_by = request.headers.get("X-Admin-User", "owner")
    log_activity(recorded_by, "CHANGE_ADMIN_PASSWORD", f"Changed password for admin: {username}")
    return jsonify({"message": f"Password for {username} updated successfully!"})

if __name__ == "__main__":
    app.run(debug=True, port=5000)
